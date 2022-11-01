import numpy as np
import os
import math
import matplotlib.pyplot as plt
import cv2
import time                                                         
from gtts import gTTS
import os

from openpyxl import Workbook
import datetime

face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
roi_gray = []

book=Workbook()
sheet=book.active

def cut_faces(image, faces_coord):                                          
    faces = []

    for (x, y, w, h) in faces_coord:                                    
        w_rm = int(0.2 * w / 2)
        faces.append(image[y : y + h, x + w_rm :  x + w - w_rm])
        
    return faces

# Load present date and time
now= datetime.datetime.now()
today=now.day
month=now.month

def add_person():
    person_name = input('What is the name of the new person: ').lower() 
    folder = 'people_folder' +'/'+ person_name                              
    
    if not os.path.exists(folder):                                  
        input("It will now take 20 pictures. Press ENTER when ready.")       
        os.mkdir(folder)                                            
        
        video = cv2.VideoCapture(0)
        detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

        counter = 1
        timer = 0
        
        cv2.namedWindow('Video Feed', cv2.WINDOW_AUTOSIZE)
        cv2.namedWindow('Saved Face', cv2.WINDOW_NORMAL)                    
        
        while counter < 21:
            _, frame = video.read()

            
            if counter == 1:
                time.sleep(6)
            else:
                time.sleep(1)
                
            faces = detector.detectMultiScale(frame)                        
            
            if len(faces):                                              
                cut_face = cut_faces(frame, faces)                          
                face_bw = cv2.cvtColor(cut_face[0], cv2.COLOR_BGR2GRAY)
                
                face_bw_eq = cv2.equalizeHist(face_bw)                      
                face_bw_eq = cv2.resize(face_bw_eq, (100, 100), interpolation = cv2.INTER_CUBIC)
                #cv2.imshow('Face Recogniser', face_bw_eq)


                cv2.imwrite(folder + '/' + str(counter) + '.jpg',
                            face_bw_eq)
                print('Images Saved:' + str(counter))
                counter += 1
                cv2.imshow('Saved Face', face_bw_eq)                        

            cv2.imshow('Video Feed', frame)
            cv2.waitKey(50)

    else:
        print("This name already exists.")                                  



def live():
    
    cv2.namedWindow('Predicting for')
    images = []
    labels = []
    labels_dic = {}
    people = [person for person in os.listdir("people_folder")]
    threshold = 105                                                     


    for i, person in enumerate(people):
        print(person)
        sheet.cell(row=1, column=int(1)).value=person
        sheet.cell(row=1, column=int(today+1)).value = "Present"
        labels_dic[i] = person
        
        for image in os.listdir("people_folder/" + person):
            images.append(cv2.imread('people_folder/'+person+'/'+ image, 0))
            labels.append(i)
            
    labels = np.array(labels)
    
    #rec_eig = cv2.face.EigenFaceRecognizer_create()
    rec_lbhp = cv2.face.LBPHFaceRecognizer_create()
    
    rec_lbhp.train(images, labels)
    
    cv2.namedWindow('face')
    webcam = cv2.VideoCapture(0)
    while True:
        _, frame = webcam.read()
        
        faces = face_cascade.detectMultiScale(frame, 1.3, 5)
        
        if len(faces):
            cut_face = cut_faces(frame, faces)                            
            
            face = cv2.cvtColor(cut_face[0], cv2.COLOR_BGR2GRAY)
            face = cv2.equalizeHist(face)                                   
            face = cv2.resize(face, (100, 100), interpolation = cv2.INTER_CUBIC)
            
            cv2.imshow('face', face)
                              
            collector = cv2.face.StandardCollector_create()
            rec_lbhp.predict_collect(face, collector)
            conf = collector.getMinDist()                            
                              
            print('Confidence ', conf)
            pred = collector.getMinLabel()
            txt = ''
            
            if conf < threshold: 
                txt = labels_dic[pred].upper()
            else:
                txt = 'Unknown'                                          
                              
            cv2.putText(frame, txt, (faces[0][0], faces[0][1] - 10), cv2.FONT_HERSHEY_PLAIN, 3, (66, 53, 243), 2)
                              
            print(faces)
            cv2.rectangle(frame, (faces[0][0], faces[0][1]),(faces[0][0] + faces[0][2], faces[0][1] + faces[0][3]), (255, 255, 0), 8)
                              
            cv2.putText(frame,"ESC to exit", (5, frame.shape[0] - 10),
                        cv2.FONT_HERSHEY_PLAIN, 1.3, (66, 53, 243), 2, cv2.LINE_AA)

        cv2.imshow("Live", frame)

        # Save Woorksheet as present month
        book.save(str(month)+'.xlsx')

        if cv2.waitKey(20) & 0xFF == 27:
            cv2.destroyAllWindows()
            break


while True:
    print("Hello there please select one of the below")
    print('Press 1 for adding a new face')
    print('Press 2 for the live recognition')
    print('Press 3 to exit')

    choice = int(input())


    if choice > 3 or choice < 1:
        print('Please select a valid choice')
    if choice == 1:
        add_person()
    elif choice == 2:
        live()
    elif choice == 3:
        print('You opted to exit!')
        break

    cv2.destroyAllWindows()
